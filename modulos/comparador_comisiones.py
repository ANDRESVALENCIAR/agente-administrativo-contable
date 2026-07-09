"""
Comparador de comisiones VIA × Contabilidad.
Lógica equivalente al HTML comparador (SheetJS → pandas/openpyxl).
"""
from __future__ import annotations

import json
import logging
import os
import re
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from typing import Any

import pandas as pd

from config import cfg  # noqa: F401 — reservado
from database import get_conn, registrar_accion, registrar_documento

# v1.1 — export Excel bytes

logger = logging.getLogger(__name__)

SFE_REGEX = re.compile(r"SFE\s*(\d{4,6})", re.I)


def _proyecto_root() -> Path:
    return Path(__file__).resolve().parent.parent


def carpeta_comisiones() -> Path:
    base = _proyecto_root() / "documentos" / "comisiones"
    base.mkdir(parents=True, exist_ok=True)
    return base


def carpeta_periodo(periodo: str) -> Path:
    p = carpeta_comisiones() / "consolidado" / periodo
    p.mkdir(parents=True, exist_ok=True)
    return p


def inicializar_tablas_comparador() -> None:
    conn = get_conn()
    c = conn.cursor()
    c.executescript("""
    CREATE TABLE IF NOT EXISTS comisiones_cruce_historial (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        titulo TEXT NOT NULL,
        periodo TEXT NOT NULL,
        fecha_cruce TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        via_filas INTEGER DEFAULT 0,
        cont_filas INTEGER DEFAULT 0,
        coincidencias INTEGER DEFAULT 0,
        solo_via INTEGER DEFAULT 0,
        solo_cont INTEGER DEFAULT 0,
        sin_sfe INTEGER DEFAULT 0,
        ruta_excel TEXT,
        ruta_consolidado TEXT,
        metricas_json TEXT
    );
    """)
    conn.commit()
    conn.close()


def _leer_hoja(path_or_buf) -> list[list]:
    df = pd.read_excel(path_or_buf, header=None, engine="openpyxl")
    return df.where(pd.notna(df), None).values.tolist()


def process_via(path_or_buf) -> dict[str, Any]:
    aoa = _leer_hoja(path_or_buf)
    if not aoa:
        raise ValueError("La hoja VIA está vacía")

    header_row_idx, sfe_col = -1, -1
    for r in range(min(len(aoa), 10)):
        for c, v in enumerate(aoa[r] or []):
            if isinstance(v, str) and re.search(r"#\s*sfe", v, re.I):
                header_row_idx, sfe_col = r, c
                break
        if header_row_idx >= 0:
            break
    if header_row_idx < 0:
        raise ValueError('No se encontró la columna con "#sfe"')

    headers_raw = [h if h is not None else f"Col_{i}" for i, h in enumerate(aoa[header_row_idx])]

    abono_cols: list[int] = []
    for c, h in enumerate(headers_raw):
        hs = str(h or "").strip()
        if re.match(r"^Abono(\.\d+)?$", hs, re.I) and len(abono_cols) < 3:
            abono_cols.append(c)
    if len(abono_cols) < 1:
        for c, h in enumerate(headers_raw):
            if re.search(r"abono", str(h or ""), re.I) and c not in abono_cols and len(abono_cols) < 3:
                abono_cols.append(c)
    if not abono_cols:
        raise ValueError("No se identificaron columnas Abono 1/2/3")

    last_keep_col = abono_cols[0] - 1
    raw_count, kept = 0, 0
    clean_rows: list[dict] = []

    for r in range(header_row_idx + 1, len(aoa)):
        row = aoa[r] or []
        if not row:
            continue
        raw_count += 1
        sfe_val = row[sfe_col] if sfe_col < len(row) else None
        if sfe_val is None or sfe_val == "":
            continue
        try:
            sfe_num = float(sfe_val)
        except (TypeError, ValueError):
            continue
        if sfe_num == 0:
            continue

        total_abono = 0.0
        for ac in abono_cols:
            if ac < len(row) and row[ac] not in (None, "") and not pd.isna(row[ac]):
                try:
                    total_abono += float(row[ac])
                except (TypeError, ValueError):
                    pass

        clean_row: dict[str, Any] = {}
        for c in range(last_keep_col + 1):
            key = str(headers_raw[c])
            val = row[c] if c < len(row) else None
            if isinstance(val, float) and pd.isna(val):
                val = None
            clean_row[key] = val
        clean_row["Total Abono"] = total_abono
        clean_row["_SFE_KEY"] = str(int(sfe_num))
        clean_rows.append(clean_row)
        kept += 1

    clean_headers = [str(h) for h in headers_raw[: last_keep_col + 1]] + ["Total Abono"]
    unique_sfes = len({r["_SFE_KEY"] for r in clean_rows})
    sum_abono = sum(r.get("Total Abono", 0) or 0 for r in clean_rows)

    return {
        "raw_count": raw_count,
        "clean_rows": clean_rows,
        "clean_headers": clean_headers,
        "unique_sfes": unique_sfes,
        "sum_abono": sum_abono,
        "kept": kept,
    }


def process_cont(path_or_buf) -> dict[str, Any]:
    aoa = _leer_hoja(path_or_buf)
    if not aoa:
        raise ValueError("La hoja Contabilidad está vacía")

    header_row_idx = 0
    while header_row_idx < len(aoa) and (
        not aoa[header_row_idx] or all(x is None or x == "" for x in aoa[header_row_idx])
    ):
        header_row_idx += 1

    headers_raw = [h if h is not None else f"Col_{i}" for i, h in enumerate(aoa[header_row_idx])]
    nota_col = next(
        (i for i, h in enumerate(headers_raw) if re.match(r"^Nota$", str(h or "").strip(), re.I)),
        -1,
    )
    if nota_col < 0:
        raise ValueError('No se encontró la columna "Nota"')

    clean_headers = [str(h) for h in headers_raw] + ["SFE"]
    raw_count = totals_removed = with_sfe = without_sfe = 0
    clean_rows: list[dict] = []

    for r in range(header_row_idx + 1, len(aoa)):
        row = aoa[r] or []
        if not row or all(x is None or x == "" for x in row):
            continue
        raw_count += 1

        if any(isinstance(c, str) and re.search(r"total", c, re.I) for c in row):
            totals_removed += 1
            continue

        nota = row[nota_col] if nota_col < len(row) else None
        m = SFE_REGEX.search(str(nota)) if nota else None
        sfe = m.group(1) if m else ""
        if sfe:
            with_sfe += 1
        else:
            without_sfe += 1

        clean_row = {str(headers_raw[i]): (row[i] if i < len(row) else None) for i in range(len(headers_raw))}
        clean_row["SFE"] = sfe
        clean_row["_SFE_KEY"] = sfe
        clean_rows.append(clean_row)

    return {
        "clean_rows": clean_rows,
        "clean_headers": clean_headers,
        "raw_count": raw_count,
        "totals_removed": totals_removed,
        "with_sfe": with_sfe,
        "without_sfe": without_sfe,
    }


def cruzar(via: dict, cont: dict) -> dict[str, Any]:
    via_map: dict[str, list] = {}
    for r in via["clean_rows"]:
        via_map.setdefault(r["_SFE_KEY"], []).append(r)

    cont_map: dict[str, list] = {}
    no_sfe_rows: list[dict] = []
    for r in cont["clean_rows"]:
        k = r.get("_SFE_KEY") or ""
        if not k:
            no_sfe_rows.append(r)
            continue
        cont_map.setdefault(k, []).append(r)

    both, only_via, only_cont = [], [], []

    for sfe, via_rows in via_map.items():
        if sfe in cont_map:
            for vr in via_rows:
                for cr in cont_map[sfe]:
                    total_abono = float(vr.get("Total Abono") or 0)
                    saldo = float(cr.get("Saldo") or 0)
                    both.append({
                        "SFE": sfe,
                        "Empresa (VIA)": vr.get("(6) Empresa", ""),
                        "Tercero (Cont.)": cr.get("Tercero", ""),
                        "Fecha SFE": vr.get("(2) Fecha", ""),
                        "Fecha Pago": cr.get("Fecha", ""),
                        "Asesor": vr.get("(17) Asesor", ""),
                        "Vendedor (Cont.)": cr.get("Vendedor", ""),
                        "Total Abono VIA": total_abono,
                        "Saldo Contabilidad": saldo,
                        "Diferencia (Cont. - VIA)": saldo - total_abono,
                        "Cuenta": cr.get("Cuenta", ""),
                        "Nota": cr.get("Nota", ""),
                        "Doc Num": cr.get("Doc Num", ""),
                    })
        else:
            for vr in via_rows:
                only_via.append({
                    "SFE": sfe,
                    "Fecha SFE": vr.get("(2) Fecha", ""),
                    "NIT": vr.get("(5) NIT", ""),
                    "Empresa": vr.get("(6) Empresa", ""),
                    "Sub. con Des. $": vr.get("(12) Sub. con Des. $", ""),
                    "Asesor": vr.get("(17) Asesor", ""),
                    "CI": vr.get("(20) CI", ""),
                    "Notas": vr.get("Notas", ""),
                    "Total Abono VIA": vr.get("Total Abono", 0),
                })

    for sfe, cont_rows in cont_map.items():
        if sfe not in via_map:
            for cr in cont_rows:
                only_cont.append({
                    "SFE": sfe,
                    "Vendedor": cr.get("Vendedor", ""),
                    "Tercero": cr.get("Tercero", ""),
                    "Cuenta": cr.get("Cuenta", ""),
                    "Fecha Pago": cr.get("Fecha", ""),
                    "Nota": cr.get("Nota", ""),
                    "Doc Num": cr.get("Doc Num", ""),
                    "Saldo": cr.get("Saldo", 0),
                })

    no_sfe = [
        {
            "Vendedor": cr.get("Vendedor", ""),
            "Tercero": cr.get("Tercero", ""),
            "Cuenta": cr.get("Cuenta", ""),
            "Fecha Pago": cr.get("Fecha", ""),
            "Nota": cr.get("Nota", ""),
            "Doc Num": cr.get("Doc Num", ""),
            "Saldo": cr.get("Saldo", 0),
        }
        for cr in no_sfe_rows
    ]

    return {"both": both, "onlyVia": only_via, "onlyCont": only_cont, "noSfe": no_sfe}


def _df_via(via: dict) -> pd.DataFrame:
    rows = [{h: r.get(h) for h in via["clean_headers"]} for r in via["clean_rows"]]
    return pd.DataFrame(rows)


def _df_cont(cont: dict) -> pd.DataFrame:
    rows = [{h: r.get(h) for h in cont["clean_headers"]} for r in cont["clean_rows"]]
    return pd.DataFrame(rows)


def _resumen_rows(via: dict, cont: dict, cross: dict, titulo: str, periodo: str) -> list[dict]:
    return [
        {"Sección": "Procesamiento", "Métrica": "Título", "Valor": titulo or "—"},
        {"Sección": "Procesamiento", "Métrica": "Periodo", "Valor": periodo},
        {"Sección": "Procesamiento", "Métrica": "Fecha de cruce", "Valor": datetime.now().strftime("%Y-%m-%d %H:%M:%S")},
        {"Sección": "VIA", "Métrica": "Filas crudas", "Valor": via["raw_count"]},
        {"Sección": "VIA", "Métrica": "Filas con SFE válida", "Valor": via["kept"]},
        {"Sección": "VIA", "Métrica": "SFEs únicas", "Valor": via["unique_sfes"]},
        {"Sección": "VIA", "Métrica": "Σ Total Abono", "Valor": via["sum_abono"]},
        {"Sección": "Contabilidad", "Métrica": "Filas crudas", "Valor": cont["raw_count"]},
        {"Sección": "Contabilidad", "Métrica": "Filas Total eliminadas", "Valor": cont["totals_removed"]},
        {"Sección": "Contabilidad", "Métrica": "Con SFE", "Valor": cont["with_sfe"]},
        {"Sección": "Contabilidad", "Métrica": "Sin SFE", "Valor": cont["without_sfe"]},
        {"Sección": "Cruce", "Métrica": "Coincidencias (en ambos)", "Valor": len(cross["both"])},
        {"Sección": "Cruce", "Métrica": "Solo VIA", "Valor": len(cross["onlyVia"])},
        {"Sección": "Cruce", "Métrica": "Solo Contabilidad", "Valor": len(cross["onlyCont"])},
        {"Sección": "Cruce", "Métrica": "Sin SFE", "Valor": len(cross["noSfe"])},
        {"Sección": "Cruce", "Métrica": "Σ Saldo coincidente", "Valor": sum(float(r.get("Saldo Contabilidad") or 0) for r in cross["both"])},
        {"Sección": "Cruce", "Métrica": "Σ Total Abono coincidente", "Valor": sum(float(r.get("Total Abono VIA") or 0) for r in cross["both"])},
    ]


def _nombre_archivo(titulo: str, sufijo: str = "Informe") -> str:
    safe = re.sub(r"[^\w\d\sáéíóúñÑ-]", "", titulo or "Cruce_Comisiones").strip().replace(" ", "_")
    fecha = date.today().isoformat()
    return f"{safe}_{sufijo}_{fecha}.xlsx"


def _escribir_informe_excel(writer, via: dict, cont: dict, cross: dict, titulo: str, periodo: str) -> None:
    from openpyxl.styles import Font, PatternFill

    pd.DataFrame(_resumen_rows(via, cont, cross, titulo, periodo)).to_excel(writer, sheet_name="Resumen", index=False)
    _df_via(via).to_excel(writer, sheet_name="VIA Limpio", index=False)
    _df_cont(cont).to_excel(writer, sheet_name="Contabilidad Limpia", index=False)
    pd.DataFrame(cross["both"]).to_excel(writer, sheet_name="En Ambos", index=False)
    pd.DataFrame(cross["onlyVia"]).to_excel(writer, sheet_name="Solo VIA", index=False)
    pd.DataFrame(cross["onlyCont"]).to_excel(writer, sheet_name="Solo Contabilidad", index=False)
    if cross["noSfe"]:
        pd.DataFrame(cross["noSfe"]).to_excel(writer, sheet_name="Sin SFE", index=False)

    header_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for sheet_name in writer.book.sheetnames:
        ws = writer.book[sheet_name]
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font


def generar_informe_excel_bytes(
    via: dict,
    cont: dict,
    cross: dict,
    titulo: str,
    periodo: str | None = None,
) -> tuple[bytes, str]:
    """Genera informe Excel completo en memoria para descarga directa."""
    periodo = periodo or date.today().strftime("%Y-%m")
    titulo = titulo or f"Cruce {periodo}"
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _escribir_informe_excel(writer, via, cont, cross, titulo, periodo)
    buf.seek(0)
    return buf.read(), _nombre_archivo(titulo, "Informe_Cruce")


def generar_via_limpio_bytes(via: dict, titulo: str) -> tuple[bytes, str]:
    """Excel solo con VIA limpio."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        _df_via(via).to_excel(writer, sheet_name="VIA Limpio", index=False)
    buf.seek(0)
    return buf.read(), _nombre_archivo(titulo or "VIA_Limpio", "VIA_Limpio")


def exportar_excel_cruce(
    via: dict,
    cont: dict,
    cross: dict,
    titulo: str,
    periodo: str | None = None,
) -> str:
    periodo = periodo or date.today().strftime("%Y-%m")
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r"[^\w\d\sáéíóúñÑ-]", "", titulo or "Cruce").strip().replace(" ", "_")
    path = carpeta_periodo(periodo) / f"{safe}_{ts}.xlsx"

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        _escribir_informe_excel(writer, via, cont, cross, titulo or "—", periodo)

    return str(path)


def actualizar_consolidado_mensual(
    periodo: str,
    titulo: str,
    via: dict,
    cont: dict,
    cross: dict,
) -> str:
    """Acumula todos los cruzes del mes en un solo Excel consolidado."""
    path = carpeta_periodo(periodo) / f"CONSOLIDADO_{periodo}.xlsx"

    registro = {
        "Fecha cruce": datetime.now().strftime("%Y-%m-%d %H:%M"),
        "Título": titulo,
        "Coincidencias": len(cross["both"]),
        "Solo VIA": len(cross["onlyVia"]),
        "Solo Contab.": len(cross["onlyCont"]),
        "Sin SFE": len(cross["noSfe"]),
        "Σ Abono VIA": via["sum_abono"],
    }

    if path.is_file():
        try:
            hist = pd.read_excel(path, sheet_name="Historial Cruces")
        except Exception:
            hist = pd.DataFrame()
        hist = pd.concat([hist, pd.DataFrame([registro])], ignore_index=True)
    else:
        hist = pd.DataFrame([registro])

    with pd.ExcelWriter(path, engine="openpyxl", mode="w") as writer:
        hist.to_excel(writer, sheet_name="Historial Cruces", index=False)
        _df_via(via).to_excel(writer, sheet_name="VIA Limpio", index=False)
        _df_cont(cont).to_excel(writer, sheet_name="Contabilidad Limpia", index=False)
        pd.DataFrame(cross["both"]).to_excel(writer, sheet_name="En Ambos", index=False)
        pd.DataFrame(cross["onlyVia"]).to_excel(writer, sheet_name="Solo VIA", index=False)
        pd.DataFrame(cross["onlyCont"]).to_excel(writer, sheet_name="Solo Contabilidad", index=False)
        if cross["noSfe"]:
            pd.DataFrame(cross["noSfe"]).to_excel(writer, sheet_name="Sin SFE", index=False)

        resumen_mes = query_resumen_periodo(periodo)
        pd.DataFrame(resumen_mes).to_excel(writer, sheet_name="Resumen Mes", index=False)

    return str(path)


def query_resumen_periodo(periodo: str) -> list[dict]:
    conn = get_conn()
    c = conn.cursor()
    c.execute(
        """SELECT titulo, fecha_cruce, coincidencias, solo_via, solo_cont, sin_sfe, ruta_excel
           FROM comisiones_cruce_historial WHERE periodo=? ORDER BY id""",
        (periodo,),
    )
    rows = [
        {
            "Título": r[0],
            "Fecha": r[1],
            "Coincidencias": r[2],
            "Solo VIA": r[3],
            "Solo Cont.": r[4],
            "Sin SFE": r[5],
            "Archivo": r[6],
        }
        for r in c.fetchall()
    ]
    conn.close()
    return rows


def guardar_cruce(
    titulo: str,
    via: dict,
    cont: dict,
    cross: dict,
    periodo: str | None = None,
) -> dict[str, Any]:
    inicializar_tablas_comparador()
    periodo = periodo or date.today().strftime("%Y-%m")
    titulo = titulo.strip() or f"Cruce {periodo}"

    ruta = exportar_excel_cruce(via, cont, cross, titulo, periodo)
    ruta_cons = actualizar_consolidado_mensual(periodo, titulo, via, cont, cross)

    metricas = {
        "via_raw": via["raw_count"],
        "via_kept": via["kept"],
        "via_unique": via["unique_sfes"],
        "via_sum": via["sum_abono"],
        "cont_raw": cont["raw_count"],
        "cont_totals": cont["totals_removed"],
        "both": len(cross["both"]),
        "only_via": len(cross["onlyVia"]),
        "only_cont": len(cross["onlyCont"]),
        "no_sfe": len(cross["noSfe"]),
    }

    conn = get_conn()
    c = conn.cursor()
    c.execute(
        """INSERT INTO comisiones_cruce_historial
           (titulo, periodo, via_filas, cont_filas, coincidencias, solo_via, solo_cont, sin_sfe,
            ruta_excel, ruta_consolidado, metricas_json)
           VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
        (
            titulo, periodo, via["kept"], len(cont["clean_rows"]),
            metricas["both"], metricas["only_via"], metricas["only_cont"], metricas["no_sfe"],
            ruta, ruta_cons, json.dumps(metricas, ensure_ascii=False),
        ),
    )
    conn.commit()
    conn.close()

    registrar_documento("xlsx", os.path.basename(ruta), ruta, "comisiones", titulo)
    registrar_accion("comisiones", "cruce_via_contabilidad", titulo, "EXITOSO")

    return {"ruta_excel": ruta, "ruta_consolidado": ruta_cons, "metricas": metricas, "periodo": periodo}


def listar_periodos() -> list[str]:
    inicializar_tablas_comparador()
    conn = get_conn()
    c = conn.cursor()
    c.execute("SELECT DISTINCT periodo FROM comisiones_cruce_historial ORDER BY periodo DESC")
    rows = [r[0] for r in c.fetchall()]
    conn.close()
    cons = carpeta_comisiones() / "consolidado"
    if cons.is_dir():
        for p in cons.iterdir():
            if p.is_dir() and p.name not in rows:
                rows.append(p.name)
    return sorted(set(rows), reverse=True)


def listar_historial(periodo: str | None = None) -> pd.DataFrame:
    inicializar_tablas_comparador()
    if periodo:
        q = "SELECT * FROM comisiones_cruce_historial WHERE periodo=? ORDER BY id DESC"
        params = (periodo,)
    else:
        q = "SELECT * FROM comisiones_cruce_historial ORDER BY periodo DESC, id DESC"
        params = ()
    conn = get_conn()
    df = pd.read_sql(q, conn, params=params)
    conn.close()
    return df
