"""Helpers RRHH — pandas, python-dateutil, openpyxl (sin IA)."""
import json
import os
from datetime import date, datetime, datetime
from typing import Any

import pandas as pd
from dateutil import rrule
from dateutil.relativedelta import relativedelta
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from dashboard.utils.db_helper import query_df


def dias_habiles(fecha_inicio: date, fecha_fin: date) -> int:
    """Cuenta días hábiles (lun–vie) entre dos fechas inclusive."""
    if fecha_fin < fecha_inicio:
        return 0
    inicio = datetime.combine(fecha_inicio, datetime.min.time())
    fin = datetime.combine(fecha_fin, datetime.min.time())
    regla = rrule.rrule(rrule.DAILY, dtstart=inicio, until=fin, byweekday=[0, 1, 2, 3, 4])
    return sum(1 for _ in regla)


def proxima_dotacion(fecha_entrega: date) -> date:
    """Próxima entrega = fecha_entrega + 4 meses."""
    return fecha_entrega + relativedelta(months=4)


def calcular_vacaciones(empleado: str, fecha_ingreso: date) -> dict[str, Any]:
    """Días causados, disfrutados y disponibles de vacaciones."""
    hoy = date.today()
    delta = relativedelta(hoy, fecha_ingreso)
    anos = delta.years + delta.months / 12 + delta.days / 365
    causados = round(anos * 15)

    df = query_df(
        """SELECT dias_habiles FROM novedades_rrhh
           WHERE empleado=? AND tipo='Vacaciones'""",
        (empleado,),
    )
    disfrutados = int(df["dias_habiles"].fillna(0).sum()) if not df.empty else 0

    return {
        "empleado": empleado,
        "fecha_ingreso": fecha_ingreso.isoformat(),
        "anos_trabajados": round(anos, 2),
        "dias_causados": causados,
        "dias_disfrutados": disfrutados,
        "dias_disponibles": max(causados - disfrutados, 0),
    }


def calcular_prima(empleado: str, salario: float, fecha_ingreso: date) -> dict[str, Any]:
    """Prima de servicios estimada del semestre."""
    hoy = date.today()
    meses = relativedelta(hoy, fecha_ingreso).years * 12 + relativedelta(hoy, fecha_ingreso).months
    factor = min(max(meses, 1) / 12, 1)
    prima = (salario / 2) * factor
    return {
        "empleado": empleado,
        "salario": salario,
        "prima_estimada": round(prima, 0),
        "semestre": "1 (ene-jun)" if hoy.month <= 6 else "2 (jul-dic)",
    }


def dias_en_proceso(fecha_aplicacion: date) -> int:
    return max((date.today() - fecha_aplicacion).days, 0)


def parse_fecha(val: Any, default: date | None = None) -> date | None:
    """Convierte valor de BD/Excel a date; tolera NaT, None y cadenas vacías."""
    if val is None:
        return default
    if isinstance(val, date) and not isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s or s.lower() in ("nat", "none", "nan", ""):
        return default
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        return default


def _estilo_header(ws) -> None:
    fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    font = Font(bold=True)
    borde = Border(*(Side(style="thin"),) * 4)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.border = borde


def _bordes_tabla(ws, max_row: int, max_col: int) -> None:
    borde = Border(*(Side(style="thin"),) * 4)
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).border = borde


def _ancho_columnas(ws) -> None:
    for col in ws.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 2, 40)


def _guardar_excel(headers: list, filas: list, hoja: str, prefijo: str) -> str:
    os.makedirs("documentos/exports", exist_ok=True)
    hoy = date.today()
    path = os.path.join("documentos", "exports", f"{prefijo}_{hoy.month:02d}_{hoy.year}.xlsx")
    wb = Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(headers)
    _estilo_header(ws)
    for fila in filas:
        ws.append(fila)
    _bordes_tabla(ws, ws.max_row, len(headers))
    _ancho_columnas(ws)
    wb.save(path)
    return path


def exportar_novedades_excel() -> str:
    df = query_df(
        """SELECT id, empleado, tipo, fecha_inicio, fecha_fin, dias_habiles, estado, notas
           FROM novedades_rrhh ORDER BY id"""
    )
    headers = ["ID", "Empleado", "Tipo", "Fecha Inicio", "Fecha Fin", "Días Hábiles", "Estado", "Notas"]
    filas = [list(r) for r in df.itertuples(index=False, name=None)] if not df.empty else []
    return _guardar_excel(headers, filas, "Novedades", "novedades")


def exportar_dotacion_excel() -> str:
    df = query_df(
        """SELECT id, empleado, item, talla, fecha_entrega, proxima_entrega, entregado
           FROM dotacion ORDER BY proxima_entrega"""
    )
    headers = ["ID", "Empleado", "Ítem", "Talla", "Fecha Entrega", "Próxima Entrega", "Entregado"]
    filas = []
    for r in df.itertuples(index=False):
        filas.append([*r[:-1], "Sí" if r[-1] else "No"])
    return _guardar_excel(headers, filas, "Dotación", "dotacion")


def exportar_candidatos_excel() -> str:
    df = query_df(
        """SELECT id, nombre, cargo, fecha_aplicacion, estado, dias_proceso, notas
           FROM candidatos ORDER BY fecha_aplicacion DESC"""
    )
    headers = ["ID", "Nombre", "Cargo", "Fecha Aplicación", "Estado", "Días en Proceso", "Notas"]
    filas = [list(r) for r in df.itertuples(index=False, name=None)] if not df.empty else []
    return _guardar_excel(headers, filas, "Candidatos", "candidatos")


_ETIQUETAS_EXPEDIENTE = {
    "cedula_cc": "Cédula",
    "contrato": "Contrato",
    "examen_medico": "Examen médico",
    "arl": "ARL",
    "eps": "EPS",
    "certificacion_ban": "Cert. bancaria",
    "hoja_vida": "Hoja de vida",
    "afiliacion": "Afiliación",
}


def exportar_informe_empleados_activos() -> str:
    """Excel con todos los empleados activos indexados en carpetas."""
    df = query_df(
        """SELECT nombre_display, nombre_carpeta, total_documentos, completitud_pct,
                  docs_faltantes, ultima_actualizacion, ruta_carpeta
           FROM empleados_carpetas WHERE activo=1
           ORDER BY nombre_display"""
    )
    headers = [
        "Empleado",
        "Carpeta",
        "Documentos",
        "Completitud %",
        "Documentos faltantes",
        "Última actualización",
        "Ruta expediente",
    ]
    filas = []
    for row in df.itertuples(index=False):
        try:
            faltantes = json.loads(row[4] or "[]")
        except json.JSONDecodeError:
            faltantes = []
        falt_txt = ", ".join(_ETIQUETAS_EXPEDIENTE.get(f, f) for f in faltantes)
        filas.append([row[0], row[1], row[2], row[3], falt_txt, row[5], row[6]])
    return _guardar_excel(headers, filas, "Personal Activo", "informe_empleados_activos")


def resumen_vacaciones_todos() -> pd.DataFrame:
    emp = query_df("SELECT nombre, fecha_ingreso, salario FROM empleados WHERE activo=1")
    filas = []
    for _, row in emp.iterrows():
        fi = parse_fecha(row["fecha_ingreso"])
        if not fi:
            continue
        vac = calcular_vacaciones(row["nombre"], fi)
        pri = calcular_prima(row["nombre"], float(row.get("salario") or 0), fi)
        filas.append({**vac, "prima_estimada": pri["prima_estimada"]})
    return pd.DataFrame(filas) if filas else pd.DataFrame()


def exportar_archivo_general_fenix() -> str:
    """Exporta toda la memoria del Archivo General Fénix a Excel multi-hoja."""
    os.makedirs("documentos/exports", exist_ok=True)
    hoy = date.today()
    path = os.path.join(
        "documentos", "exports", f"archivo_general_fenix_{hoy.month:02d}_{hoy.year}.xlsx"
    )

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        query_df(
            """SELECT nombre_completo, cedula, cargo, departamento, jefe_inmediato, lugar_labor,
                      tipo_contrato, termino, fecha_ingreso, vencimiento_contrato, fecha_preaviso,
                      fecha_nacimiento, telefono, email_corporativo, email_personal, direccion,
                      barrio, ciudad, salario_ibc, salud, pension, cesantias, caja_compensacion,
                      ref_emergencia, parentesco_emergencia, tel_emergencia, carpeta_vinculada,
                      observaciones, origen
               FROM empleados_fenix WHERE tipo_plantilla='nomina_activo' AND activo=1
               ORDER BY nombre_completo"""
        ).to_excel(writer, sheet_name="NOMINA ACTIVA", index=False)

        query_df(
            """SELECT nombre_completo, cedula, cargo, departamento, fecha_ingreso, telefono,
                      email_corporativo, salario_ibc, tipo_contrato, carpeta_vinculada, observaciones
               FROM empleados_fenix WHERE tipo_plantilla='prestacion_servicios' AND activo=1
               ORDER BY nombre_completo"""
        ).to_excel(writer, sheet_name="PRESTACION SERVICIOS", index=False)

        query_df(
            """SELECT nombre_completo, cedula, cargo, fecha_ingreso, observaciones
               FROM empleados_fenix WHERE tipo_plantilla='retirado' OR activo=0
               ORDER BY nombre_completo"""
        ).to_excel(writer, sheet_name="RETIRADOS", index=False)

        query_df(
            """SELECT nombre_completo, cedula, fecha_ingreso, pendiente_dotacion,
                      examenes_periodicos, meses_json FROM novedades_fenix ORDER BY nombre_completo"""
        ).to_excel(writer, sheet_name="NOVEDADES", index=False)

        query_df(
            """SELECT nombre_completo, hoja_origen, dias_pendientes, dias_tomar,
                      dias_pendientes_2025, fecha_regreso, observaciones
               FROM vacaciones_fenix ORDER BY nombre_completo"""
        ).to_excel(writer, sheet_name="VACACIONES", index=False)

        query_df("SELECT nombre, dia, mes FROM cumpleanos_fenix ORDER BY mes, dia").to_excel(
            writer, sheet_name="CUMPLEANOS", index=False
        )

        query_df(
            """SELECT nombre_completo, cedula, cargo, fecha_inicio, termino,
                      vencimiento_contrato, fecha_preaviso FROM contratos_fijos_fenix
               ORDER BY nombre_completo"""
        ).to_excel(writer, sheet_name="CONTRATOS FIJOS", index=False)

    return path


def exportar_contratos_activos_excel() -> str:
    """Exporta planilla Contratos Activos con formato tipo informe horizontal."""
    from modulos.contratos_activos import COLUMNAS_INFORME

    df = query_df("SELECT * FROM contratos_activos ORDER BY nombre_completo")
    os.makedirs("documentos/exports", exist_ok=True)
    hoy = date.today()
    path = os.path.join("documentos", "exports", f"contratos_activos_{hoy.month:02d}_{hoy.year}.xlsx")

    cols = [k for k, _ in COLUMNAS_INFORME if k in df.columns]
    labels = [v for k, v in COLUMNAS_INFORME if k in df.columns]
    export = df[cols].copy() if not df.empty else pd.DataFrame(columns=cols)
    export.columns = labels

    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        export.to_excel(writer, sheet_name="CONTRATOS ACTIVOS", index=False)
        wb = writer.book
        ws = wb["CONTRATOS ACTIVOS"]
        from openpyxl.styles import Font, PatternFill

        header_fill = PatternFill(start_color="4A4A4A", end_color="4A4A4A", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=9)
        data_fill = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
        ok_fill = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
        no_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")

        doc_labels = {
            "FOTO", "CV", "ANTECEDENTES", "CONTRATO FIRMADO", "DNI ESCANEADO",
            "RECIBO LUZ/AGUA", "CROQUIS", "DECLARACIÓN JURADA", "CERTIFICADOS",
            "SCTR", "VIDA LEY", "EXAMEN MÉDICO", "INDUCCIÓN", "EPP",
        }
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                hdr = ws.cell(row=1, column=cell.column).value
                val = str(cell.value or "").upper()
                if hdr in doc_labels:
                    if val == "SI":
                        cell.fill = ok_fill
                    elif val == "NO":
                        cell.fill = no_fill
                    else:
                        cell.fill = data_fill
                else:
                    cell.fill = data_fill

    return path
