"""Generación de reportes Excel, PDF y Word — usa librerías del core."""
import os
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from config import cfg
from core.documentos_engine import generar_pdf_tabla


def _ruta_reporte(nombre: str, ext: str) -> str:
    carpeta = os.path.join("reportes", datetime.now().strftime("%Y-%m"))
    os.makedirs(carpeta, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    return os.path.join(carpeta, f"{nombre}_{ts}.{ext}")


def generar_reporte(tipo: str, datos: pd.DataFrame, formato: str) -> tuple[bytes, str]:
    """Genera reporte en excel o pdf."""
    if formato == "excel":
        return _excel(tipo, datos), f"{tipo}.xlsx"
    if formato == "pdf":
        return _pdf(tipo, datos), f"{tipo}.pdf"
    raise ValueError(f"Formato no soportado: {formato}")


def _excel(tipo: str, datos: pd.DataFrame) -> bytes:
    """XLSX con openpyxl o xlsxwriter si está instalado."""
    try:
        import xlsxwriter
        from core.registro_libs import registrar_uso_libreria

        registrar_uso_libreria("contabilidad", "xlsxwriter")
        buf = BytesIO()
        with pd.ExcelWriter(buf, engine="xlsxwriter") as writer:
            datos.to_excel(writer, sheet_name=tipo[:31], index=False)
            wb = writer.book
            ws = writer.sheets[tipo[:31]]
            fmt = wb.add_format({"bold": True, "bg_color": "#185FA5", "font_color": "white"})
            for col, name in enumerate(datos.columns):
                ws.write(0, col, name, fmt)
        path = _ruta_reporte(tipo, "xlsx")
        data = buf.getvalue()
        with open(path, "wb") as f:
            f.write(data)
        return data
    except ImportError:
        pass

    wb = Workbook()
    ws = wb.active
    ws.title = tipo[:31]
    header_fill = PatternFill(start_color="185FA5", end_color="185FA5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col_idx, col in enumerate(datos.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col)
        cell.fill = header_fill
        cell.font = header_font
    for row_idx, row in enumerate(datos.itertuples(index=False), 2):
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
    path = _ruta_reporte(tipo, "xlsx")
    buf = BytesIO()
    wb.save(buf)
    wb.save(path)
    return buf.getvalue()


def _pdf(tipo: str, datos: pd.DataFrame) -> bytes:
    """PDF vía reportlab (core) o fpdf2."""
    path = _ruta_reporte(tipo, "pdf")
    titulo = f"{cfg.NOMBRE_EMPRESA} — {tipo}"
    if datos.empty:
        cols, filas = ["info"], [["Sin datos"]]
    else:
        cols = list(datos.columns)
        filas = datos.astype(str).values.tolist()
    return generar_pdf_tabla(titulo, cols, filas, path)
